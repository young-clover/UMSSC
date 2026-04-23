import pandas as pd
from pathlib import Path
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# --- 解析逻辑 ---

def parse_instance_info(filename):
    name = Path(filename).stem
    pattern = r'([A-E])__([0-9.]+)_(SOCP|PBD_single|PBD_multi)'
    match = re.match(pattern, name, re.IGNORECASE)
    if match:
        series = match.group(1).upper()
        func_type_raw = match.group(2)
        func_type = "mixed" if func_type_raw == "0" else func_type_raw
        
        method_raw = match.group(3)
        method_map = {'PBD_multi': 'PBD-multi', 'PBD_single': 'PBD-single', 'SOCP': 'MISOCP'}
        method = method_map.get(method_raw, method_raw)
        return method, func_type, series
    return 'Unknown', 'Unknown', 'Unknown'

def parse_pbd_instance(instance_content):
    result = {}
    matches = re.finditer(r'([A-Za-z0-9_]+):([A-Za-z0-9.-]+)', instance_content)
    raw_data = {m.group(1): m.group(2) for m in matches}
    mapping = {
        'RootLowerBound': 'RootLowerBound', 'RootUpperBound': 'RootUpperBound',
        'RootGap': 'RootGap', 'RootRuntime': 'RootRuntime',
        'NumCuts': 'NumCuts', 'NumIters': 'NumIters',
        'HeurUpperBound': 'HeurUpperBound', 'HeurGap': 'HeurGap',
        'HeurRuntime': 'HeurRuntime', 'FixingRuntime': 'FixingRuntime',
        'FixedNum': 'FixedNum', 'LowerBound': 'FinalLowerBound',
        'UpperBound': 'FinalUpperBound', 'Gap': 'FinalGap',
        'BranchTime': 'BranchTime', 'Nodes': 'Nodes', 'Status': 'Status'
    }
    for k, v in raw_data.items():
        if k in mapping: result[mapping[k]] = v
    return result

def parse_socp_instance(instance_content):
    result = {}
    matches = re.finditer(r'([A-Za-z0-9_]+):([A-Za-z0-9.-]+)', instance_content)
    raw_data = {m.group(1): m.group(2) for m in matches}
    mapping = {
        'RootLowerBound': 'RootLowerBound', 'RootUpperBound': 'RootUpperBound',
        'RootGap': 'RootGap', 'RootRuntime': 'RootRuntime',
        'LowerBound': 'FinalLowerBound', 'UpperBound': 'FinalUpperBound',
        'Gap': 'FinalGap', 'BranchTime': 'BranchTime',
        'Nodes': 'Nodes', 'Status': 'Status'
    }
    for k, v in raw_data.items():
        if k in mapping: result[mapping[k]] = v
    return result

def parse_single_result_file(file_path):
    method, func_type, series = parse_instance_info(file_path.name)
    instances = []
    TIME_LIMIT_THRESHOLD = 999.0
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        instance_blocks = re.split(r'Instance:\s*([^;]+);', content)
        for i in range(1, len(instance_blocks), 2):
            if i + 1 >= len(instance_blocks): break
            instance_id = instance_blocks[i].replace('_bounds.txt', '').strip()
            instance_content = instance_blocks[i + 1]
            result = {'Series': series, 'InstanceID': instance_id, 'Method': method, 'FunctionType': func_type}
            parsed_data = parse_pbd_instance(instance_content) if 'PBD' in method else parse_socp_instance(instance_content)
            result.update(parsed_data)
            
            for field in ['RootGap', 'RootRuntime', 'HeurGap', 'HeurRuntime', 'FixingRuntime', 'FinalGap', 'BranchTime', 'Nodes']:
                if field in result and result[field] is not None:
                    try:
                        val_str = str(result[field]).strip().upper()
                        if 'Gap' in field and val_str == 'INF':
                            result[field] = 100.0
                        else:
                            val = float(result[field])
                            if 'Gap' in field:
                                result[field] = 100.0 if (val >= 1e6 or np.isinf(val)) else val * 100.0
                            else:
                                result[field] = val
                    except:
                        if 'Gap' in field: result[field] = 100.0
            
            current_gap = result.get('RootGap', 100.0)
            if (result.get('HeurRuntime') or 0) > 0:
                current_gap = result.get('HeurGap', current_gap)
            if (result.get('BranchTime') or 0) > 0 or (result.get('Nodes') or 0) > 1:
                current_gap = result.get('FinalGap', current_gap)
            result['FinalGap'] = current_gap
            
            total_time = sum(result.get(f, 0) or 0 for f in ['RootRuntime', 'HeurRuntime', 'FixingRuntime', 'BranchTime'])
            result['TotalTime'] = total_time
            
            if result.get('BranchTime', 0) <= 0 and result.get('Nodes', 0) <= 1:
                result['Nodes'] = 0
                result['Status'] = 9 if total_time >= TIME_LIMIT_THRESHOLD else 2
            else:
                try: result['Status'] = int(float(result.get('Status', 2)))
                except: result['Status'] = 2
            instances.append(result)
    except Exception as e:
        print(f"解析出错 {file_path}: {e}")
    return instances

def format_excel_sheet(ws, methods_count=3):
    """格式化：包含整数列转换、算例名合并与三线表格式"""
    is_series_sheet = ws.title.startswith("Series_")
    
    # 定义需要显示为整数的列名
    int_cols = {'#cuts', '#iters', '#fixed', '#nodes', 'status'}
    
    # 确定表头行数
    header_row = 2 if ("Stats" in ws.title) else 1
    
    # 获取每一列对应的底层表头名称
    col_headers = {}
    for col in range(1, ws.max_column + 1):
        col_headers[col] = str(ws.cell(row=header_row, column=col).value)

    # 定义科技论文三线表线条样式（顶部底部粗线，表头下方细线）
    thick_line = Side(style='medium', color='000000')
    thin_line = Side(style='thin', color='000000')

    # 1. 基础样式与三线表边框
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 字体加粗逻辑：仅表头加粗，其余全部正常显示 (除了汇总表自带的 Total Summary)
            is_total_summary = str(ws.cell(row=cell.row, column=1).value) == "Total Summary"
            if cell.row <= header_row:
                cell.font = Font(bold=True)
            elif not is_series_sheet and is_total_summary:
                cell.font = Font(bold=True, color="FF0000") # 保留总表的标红汇总
            else:
                cell.font = Font(bold=False) # 强制常规字体
            
            # 数值格式处理
            if cell.row > header_row and isinstance(cell.value, (int, float)):
                header_name = col_headers.get(cell.column, "")
                if header_name in int_cols:
                    cell.number_format = '0' # 整数格式
                else:
                    cell.number_format = '0.00' # 两位小数格式

            # 三线表边框绘制逻辑
            border_top = None
            border_bottom = None
            
            if cell.row == 1:
                border_top = thick_line # 顶线 (粗)
            if cell.row == header_row:
                border_bottom = thin_line # 表头底线 (细)
            if cell.row == ws.max_row:
                border_bottom = thick_line # 底线 (粗)
                
            cell.border = Border(top=border_top, bottom=border_bottom, left=None, right=None)

    # 2. 合并第一列（InstanceID）
    if is_series_sheet:
        start_row = header_row + 1 
        max_row = ws.max_row
        for r in range(start_row, max_row + 1, methods_count):
            if r + methods_count - 1 <= max_row:
                ws.merge_cells(start_row=r, start_column=1, end_row=r + methods_count - 1, end_column=1)
                ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # 3. 统一调整列宽
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

# --- 主流程 ---

def extract_all_results(results_dir, output_excel):
    results_dir = Path(results_dir)
    files = list(results_dir.glob('*.txt')) + list(results_dir.glob('*.log'))
    all_data = []
    for f in files: all_data.extend(parse_single_result_file(f))
    if not all_data: return
    df = pd.DataFrame(all_data)
    
    methods_order = ['PBD-multi', 'PBD-single', 'MISOCP']
    num_methods = len(methods_order)
    
    stats_rename_map = {
        'RootGap': '$g_r$[%]', 'HeurGap': '$g_h$[%]', 'FinalGap': '$g$[%]',
        'RootRuntime': '$t_r$[s]', 'HeurRuntime': '$t_h$[s]', 'TotalTime': '$t$[s]',
        'Status': 'status'
    }
    
    series_rename_map = {
        'RootLowerBound': '$lb_r$', 'RootUpperBound': '$ub_r$', 'RootGap': '$g_r$[%]', 'RootRuntime': '$t_r$[s]',
        'NumCuts': '#cuts', 'NumIters': '#iters', 'HeurUpperBound': '$ub_h$', 'HeurGap': '$g_h$[%]',
        'HeurRuntime': '$t_h$[s]', 'FixingRuntime': '$t_f$[s]', 'FixedNum': '#fixed',
        'FinalLowerBound': '$lb$', 'FinalUpperBound': '$ub$', 'FinalGap': '$g$[%]',
        'BranchTime': '$t_b$[s]', 'Nodes': '#nodes', 'Status': 'status', 'TotalTime': '$t$[s]'
    }

    # 1. Stats 表处理
    avg_metrics = ['RootGap', 'HeurGap', 'FinalGap', 'RootRuntime', 'HeurRuntime', 'TotalTime']
    df_avg = df.groupby(['InstanceID', 'Method'])[avg_metrics].mean().reset_index()
    df_status = df.groupby(['InstanceID', 'Method'])['Status'].apply(lambda x: f"{(x==2).sum()}/{len(x)}").reset_index()
    df_merged = pd.merge(df_avg, df_status, on=['InstanceID', 'Method'])
    df_pivot = df_merged.pivot(index='InstanceID', columns='Method').swaplevel(0, 1, axis=1)
    
    df_func_avg = df.groupby(['FunctionType', 'Method'])[['RootGap', 'HeurGap', 'FinalGap']].mean().reset_index()
    df_func_status = df.groupby(['FunctionType', 'Method'])['Status'].apply(lambda x: f"{(x==2).sum()}/{len(x)}").reset_index()
    df_func_merged = pd.merge(df_func_avg, df_func_status, on=['FunctionType', 'Method'])
    df_func_pivot = df_func_merged.pivot(index='FunctionType', columns='Method').swaplevel(0, 1, axis=1)

    def add_summary_row(pivot_df):
        summary = pd.Series(index=pivot_df.columns, name='Total Summary', dtype='object')
        for col in pivot_df.columns:
            metric = col[1]
            if 'Gap' in metric:
                summary[col] = pivot_df[col].mean()
            elif metric == 'Status':
                total_s, total_n = 0, 0
                for val in pivot_df[col].dropna():
                    try:
                        s, n = map(int, str(val).split('/'))
                        total_s += s; total_n += n
                    except: pass
                summary[col] = f"{total_s}/{total_n}"
            else:
                summary[col] = np.nan
        return pd.concat([pivot_df, pd.DataFrame([summary])])

    def process_pivot_data(pivot_df, is_func=False):
        final_cols = []
        for m in methods_order:
            if m in pivot_df.columns.levels[0]:
                if is_func:
                    sub = (['RootGap', 'HeurGap', 'FinalGap', 'Status'] if 'PBD' in m else ['RootGap', 'FinalGap', 'Status'])
                else:
                    sub = (['RootGap', 'HeurGap', 'FinalGap', 'RootRuntime', 'HeurRuntime', 'TotalTime', 'Status'] if 'PBD' in m 
                           else ['RootGap', 'FinalGap', 'RootRuntime', 'TotalTime', 'Status'])
                for sc in sub:
                    if (m, sc) in pivot_df.columns: final_cols.append((m, sc))
        pivot_df = pivot_df[final_cols]
        pivot_df = add_summary_row(pivot_df)
        pivot_df.columns = pivot_df.columns.map(lambda x: (x[0], stats_rename_map.get(x[1], x[1])))
        return pivot_df

    df_pivot_final = process_pivot_data(df_pivot, False)
    desired_func_order = ['1.5', '2.0', '2.5', '3.0', 'mixed']
    df_func_pivot_final = process_pivot_data(df_func_pivot.reindex([f for f in desired_func_order if f in df_func_pivot.index]), True)

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df_pivot_final.to_excel(writer, sheet_name='Average Stats')
        df_func_pivot_final.to_excel(writer, sheet_name='Function Stats')

        # 2. 子表处理
        INF_THRESHOLD = 1e20
        present_funcs = [f for f in desired_func_order if f in df['FunctionType'].unique()]
        for ft in present_funcs:
            for s in ['A', 'B', 'C', 'D', 'E']:
                s_df = df[(df['Series'] == s) & (df['FunctionType'] == ft)].copy()
                if s_df.empty: continue
                
                raw_cols = [c for c in series_rename_map.keys() if c in s_df.columns]
                for c in raw_cols:
                    if c != 'Status': s_df[c] = pd.to_numeric(s_df[c], errors='coerce')
                
                agg_funcs = {c: ('first' if c == 'Status' else 'mean') for c in raw_cols}
                s_grouped = s_df.groupby(['InstanceID', 'Method'])[raw_cols].agg(agg_funcs)
                
                all_instances = sorted(s_df['InstanceID'].unique())
                new_index = pd.MultiIndex.from_product([all_instances, methods_order], names=['InstanceID', 'Method'])
                s_reindexed = s_grouped.reindex(new_index).reset_index()
                
                # --- 新增/修改：统一处理所有 ub 相关的极大值替换 ---
                ub_related_cols = ['RootUpperBound', 'FinalUpperBound']
                for col in ub_related_cols:
                    if col in s_reindexed.columns:
                        s_reindexed[col] = s_reindexed[col].apply(
                            lambda x: 'INF' if pd.notnull(x) and isinstance(x, (float, int)) and (x >= INF_THRESHOLD or np.isinf(x)) else x
                        )
                
                s_reindexed = s_reindexed.rename(columns={'InstanceID': f'p = {ft}'})
                s_reindexed = s_reindexed.drop(columns=['FunctionType'], errors='ignore')
                s_reindexed = s_reindexed.rename(columns=series_rename_map)
                
                s_final = s_reindexed.fillna('------')
                sheet_name = f'Series_{s}_{ft}'[:31]
                s_final.to_excel(writer, sheet_name=sheet_name, index=False)

    # 3. 后处理格式
    wb = load_workbook(output_excel)
    for sn in wb.sheetnames: format_excel_sheet(wb[sn], methods_count=num_methods)
    wb.save(output_excel)
    print("✓ 处理完成！")

if __name__ == "__main__":
    current_dir = Path(__file__).parent.resolve()
    extract_all_results(current_dir, current_dir / "results_final_summary.xlsx")